import re, zipfile, io, json
from pathlib import Path
from bs4 import BeautifulSoup

# ---- CONFIG ----
ZIP_FILE = "submissions (breakeven).zip"   # <-- change to your actual zip filename
OUT_FILE = "urls.txt"
# Accept both chat.openai.com and chatgpt.com share links (with optional query parts)
PATTERN = re.compile(r"https?://(chat\.openai\.com|chatgpt\.com)/share/[A-Za-z0-9\-_]+(?:\?[^\s\"'<>]*)?", re.I)

# Optional: list what we found per file
VERBOSE = True

def try_decode(data: bytes):
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(enc, errors="ignore")
        except Exception:
            continue
    return ""

def extract_from_html(text: str):
    urls = set(PATTERN.findall(text))  # Find domain groups only; fix below
    # Correct the above: re.findall returns tuples when groups present, so re-run plain finditer:
    return set(m.group(0) for m in PATTERN.finditer(text))

def extract_from_html_bs(html_text: str):
    urls = set()
    soup = BeautifulSoup(html_text, "lxml")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if PATTERN.search(href):
            urls.add(href)
    # Also search the raw HTML text
    urls |= extract_from_html(html_text)
    return urls

def extract_from_docx_bytes(data: bytes):
    # Lazy import to avoid dependency errors if not installed
    try:
        from docx import Document
    except ImportError:
        return set()
    urls = set()
    f = io.BytesIO(data)
    doc = Document(f)
    # Text content
    for p in doc.paragraphs:
        t = p.text or ""
        urls |= extract_from_html(t)
    # Hyperlinks in relationships
    try:
        rels = doc.part.rels.values()
        for r in rels:
            target = getattr(r._target, "ref", None) or getattr(r._target, "partname", None) or getattr(r._target, "external_reference", None)
        # Better: iterate rels and use _target if has 'ref' attr or to string
        for r in doc.part.rels.values():
            try:
                target = str(r._target)  # often a URL string
                if PATTERN.search(target):
                    urls.add(target)
            except Exception:
                pass
    except Exception:
        pass
    return urls

def extract_from_xlsx_bytes(data: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return set()
    urls = set()
    f = io.BytesIO(data)
    wb = load_workbook(f, data_only=True, read_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    if PATTERN.search(cell):
                        urls |= extract_from_html(cell)
    return urls

def extract_from_pdf_bytes(data: bytes):
    try:
        from pypdf import PdfReader
    except ImportError:
        return set()
    urls = set()
    try:
        reader = PdfReader(io.BytesIO(data))
        # Text
        for page in reader.pages:
            try:
                txt = page.extract_text() or ""
                if txt:
                    urls |= extract_from_html(txt)
            except Exception:
                pass
        # Annotations / links
        for page in reader.pages:
            annots = page.get("/Annots", [])
            for a in annots or []:
                try:
                    obj = a.get_object()
                    if "/A" in obj and "/URI" in obj["/A"]:
                        href = obj["/A"]["/URI"]
                        if isinstance(href, str) and PATTERN.search(href):
                            urls.add(href)
                except Exception:
                    pass
    except Exception:
        pass
    return urls

def extract_from_json_text(text: str):
    urls = set()
    try:
        # quick and dirty: regex over raw text first
        urls |= extract_from_html(text)
        # if it's valid json, walk values
        obj = json.loads(text)
        def walk(v):
            if isinstance(v, dict):
                for x in v.values(): walk(x)
            elif isinstance(v, list):
                for x in v: walk(x)
            elif isinstance(v, str):
                if PATTERN.search(v): urls.add(v)
        walk(obj)
    except Exception:
        pass
    return urls

def main():
    zpath = Path(ZIP_FILE)
    if not zpath.exists():
        print(f"ERROR: ZIP not found at {zpath.resolve()}")
        return

    seen = set()
    with zipfile.ZipFile(zpath, "r") as z:
        for name in z.namelist():
            # skip directories
            if name.endswith("/") or name.endswith("\\"):
                continue
            try:
                data = z.read(name)
            except Exception:
                continue

            ext = Path(name).suffix.lower()
            found = set()

            if ext in {".html", ".htm"}:
                text = try_decode(data)
                found |= extract_from_html_bs(text)
            elif ext in {".txt", ".md", ".csv"}:
                text = try_decode(data)
                found |= extract_from_html(text)
            elif ext == ".docx":
                found |= extract_from_docx_bytes(data)
            elif ext in {".xlsx", ".xlsm"}:
                found |= extract_from_xlsx_bytes(data)
            elif ext == ".pdf":
                found |= extract_from_pdf_bytes(data)
            elif ext in {".json", ".ndjson", ".jsonl"}:
                text = try_decode(data)
                found |= extract_from_json_text(text)
            elif ext in {".url"}:
                # Windows InternetShortcut files are INI-like
                text = try_decode(data)
                found |= extract_from_html(text)
            else:
                # Fallback: try plain decode + regex
                text = try_decode(data)
                if text:
                    found |= extract_from_html(text)

            if VERBOSE and found:
                print(f"[+] {name}: {len(found)} link(s)")

            seen |= found

    if not seen:
        print("No links found. Consider checking:")
        print(" - ZIP_FILE name is correct")
        print(" - Links are share links (chat.openai.com/share or chatgpt.com/share)")
        print(" - Files inside ZIP are supported types (txt, docx, xlsx, html, json, pdf)")
    Path(OUT_FILE).write_text("\n".join(sorted(seen)), encoding="utf-8")
    print(f"✅ Extracted {len(seen)} links into {OUT_FILE}")

if __name__ == "__main__":
    main()

